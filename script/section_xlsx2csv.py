import traceback
import sys
import re
import getpass

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print('openpyxl is not installed. Please install for  user "' + getpass.getuser() + '"')
    sys.exit(1)

if len(sys.argv) != 2:
    sys.exit("Usage: python " + sys.argv[0] + " input")

input_file = sys.argv[1]

try:
    wb = load_workbook(input_file)
except Exception:
    print(traceback.format_exc())
    exit(-1)

out = ""

# If you're reading this, it probably means this function finally broke.
# You can't repair it, don't even try. Just build a new one. Save yourself the sanity.
try:
    for ws in wb._sheets:
        for row in ws.iter_rows(values_only=True):
            if type(row[0]) == str or type(row[0]) == str:
                if re.match(r'\d{4} +\w{2,4} +\d{3} +\d{2} +.+', row[0]):
                    out += re.sub(r'\s+', '\t', row[0], 4) + '\n'
except Exception:
    print(traceback.format_exc())
    exit(-1)

print(out)
